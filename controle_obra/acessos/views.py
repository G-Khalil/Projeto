from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
import base64
import json
from PIL import Image
from io import BytesIO
import numpy as np
import face_recognition
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from acessos.models import AcessoObra
from empresas.models import Empresa
from funcionarios.models import Funcionario

def lista_presenca_hoje(request):
    """
    Tela de presença de hoje agrupada por empresa.
    """
    data_str = request.GET.get('data', timezone.now().strftime('%Y-%m-%d'))
    hoje = datetime.strptime(data_str, '%Y-%m-%d').date()
    dados_empresas = []
    empresas = Empresa.objects.all().order_by('nome')
    
    for empresa in empresas:
        funcionarios = Funcionario.objects.filter(empresa=empresa).order_by('nome')
        acessos_empresa = []
        for func in funcionarios:
            acesso = AcessoObra.objects.filter(funcionario=func, data=hoje).first()
            acessos_empresa.append({
                'funcionario': func,
                'acesso': acesso
            })
        
        if acessos_empresa:
            dados_empresas.append({
                'empresa': empresa,
                'dados': acessos_empresa
            })

    context = {
        'dados_empresas': dados_empresas,
        'data': hoje,
    }
    return render(request, 'acessos/lista_presenca_hoje.html', context)
def lista_funcionarios_entrada_saida(request):
    """
    Lista de funcionários agrupada por empresa com status de entrada/saída.
    """
    hoje = timezone.now().date()
    empresas = Empresa.objects.all().order_by('nome')
    lista_final = []
    
    for empresa in empresas:
        funcionarios = Funcionario.objects.filter(empresa=empresa).order_by('nome')
        funcs_status = []
        for func in funcionarios:
            acesso = AcessoObra.objects.filter(funcionario=func, data=hoje).first()
            status = 'nao-chegou'
            display = 'Não chegou'
            if acesso:
                if acesso.hora_saida:
                    status = 'saida'
                    display = 'Saída registrada'
                else:
                    status = 'entrada'
                    display = 'Entrada registrada'
            
            funcs_status.append({
                'id': func.id,
                'nome': func.nome,
                'status': status,
                'status_display': display,
                'hora_entrada': acesso.hora_entrada if acesso else None,
                'hora_saida': acesso.hora_saida if acesso else None,
                'tem_foto': bool(func.foto),
                'foto_url': func.foto.url if func.foto else None
            })
        if funcs_status:
            lista_final.append({'empresa': empresa.nome, 'funcionarios': funcs_status})
    
    # Calcular contagens
    presentes_count = 0
    ausentes_count = 0
    saidas_count = 0
    
    for empresa_data in lista_final:
        for func in empresa_data['funcionarios']:
            if func['status'] == 'entrada':
                presentes_count += 1
            elif func['status'] == 'saida':
                saidas_count += 1
            else:  # nao-chegou
                ausentes_count += 1
    
    return render(request, 'acessos/lista_funcionarios_entrada_saida.html', {'empresas_agrupadas': lista_final, 'data': hoje, 'presentes_count': presentes_count, 'ausentes_count': ausentes_count, 'saidas_count': saidas_count})


def exportar_excel_mensal(request):
    """Exporta relatório de acessos mensal em Excel."""
    mes = int(request.GET.get('mes', timezone.now().month))
    ano = int(request.GET.get('ano', timezone.now().year))
    
    # Buscar acessos do mês
    acessos = AcessoObra.objects.filter(data__month=mes, data__year=ano).select_related('funcionario')
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f'Acessos {mes:02d}/{ano}'
    ws.append(['Data', 'Funcionário', 'Hora Entrada', 'Hora Saída'])
    
    for acesso in acessos:
        ws.append([
            str(acesso.data),
            acesso.funcionario.nome,
            str(acesso.hora_entrada) if acesso.hora_entrada else '',
            str(acesso.hora_saida) if acesso.hora_saida else ''
        ])
    
    # Salvar workbook em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="acessos_{mes:02d}_{ano}.xlsx"'
    return response


def exportar_excel_diario(request):
    """Exporta relatório de acessos do dia atual em Excel."""
    hoje = timezone.now().date()
    acessos = AcessoObra.objects.filter(data=hoje).select_related('funcionario')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f'Acessos {hoje}'
    ws.append(['Data', 'Funcionário', 'Hora Entrada', 'Hora Saída'])
    for acesso in acessos:
        ws.append([str(acesso.data), acesso.funcionario.nome, str(acesso.hora_entrada) if acesso.hora_entrada else '', str(acesso.hora_saida) if acesso.hora_saida else ''])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="acessos_{hoje}.xlsx"'
    return response

